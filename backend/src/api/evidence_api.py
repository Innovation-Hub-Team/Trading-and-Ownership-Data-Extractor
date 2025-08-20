#!/usr/bin/env python3
"""
Flask API for serving evidence screenshots and extraction metadata
"""

from flask import Flask, send_file, jsonify, request
from flask_cors import CORS
import json
import os
from pathlib import Path
import logging
from datetime import datetime
import pandas as pd
from werkzeug.utils import secure_filename
import sys

# Add the src directory to the path so we can import our modules
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))

# Import the new extractor2 functionality
from extractors.extractor2 import (
    find_headings_in_pdf, 
    export_pages_to_png, 
    extract_financial_metrics_with_gemini,
    configure_gemini,
    DEFAULT_TARGET_TITLES
)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def parse_report_date(date_string: str) -> str:
    """
    Parse and standardize report date from various formats.
    Returns date in YYYY-MM-DD format.
    """
    if not date_string or date_string == "NOT_FOUND":
        return datetime.now().strftime('%Y-%m-%d')
    
    try:
        # Try to parse various date formats
        import re
        from datetime import datetime
        
        # Format 1: DD Month YYYY (e.g., "14 November 2024")
        month_pattern = r'(\d{1,2})\s+(\w+)\s+(\d{4})'
        match = re.search(month_pattern, date_string, re.IGNORECASE)
        if match:
            day, month_name, year = match.groups()
            try:
                # Convert month name to number
                month_map = {
                    'january': 1, 'jan': 1,
                    'february': 2, 'feb': 2,
                    'march': 3, 'mar': 3,
                    'april': 4, 'apr': 4,
                    'may': 5,
                    'june': 6, 'jun': 6,
                    'july': 7, 'jul': 7,
                    'august': 8, 'aug': 8,
                    'september': 9, 'sep': 9, 'sept': 9,
                    'october': 10, 'oct': 10,
                    'november': 11, 'nov': 11,
                    'december': 12, 'dec': 12
                }
                month_num = month_map.get(month_name.lower())
                if month_num:
                    return f"{year}-{month_num:02d}-{int(day):02d}"
            except:
                pass
        
        # Format 2: MM/DD/YYYY (e.g., "11/14/2024")
        slash_pattern = r'(\d{1,2})/(\d{1,2})/(\d{4})'
        match = re.search(slash_pattern, date_string)
        if match:
            month, day, year = match.groups()
            return f"{year}-{int(month):02d}-{int(day):02d}"
        
        # Format 3: YYYY-MM-DD (already correct)
        iso_pattern = r'(\d{4})-(\d{1,2})-(\d{1,2})'
        match = re.search(iso_pattern, date_string)
        if match:
            year, month, day = match.groups()
            return f"{year}-{int(month):02d}-{int(day):02d}"
        
        # Format 4: DD-MM-YYYY (e.g., "14-11-2024")
        dash_pattern = r'(\d{1,2})-(\d{1,2})-(\d{4})'
        match = re.search(dash_pattern, date_string)
        if match:
            day, month, year = match.groups()
            return f"{year}-{int(month):02d}-{int(day):02d}"
        
        # If no pattern matches, return as is (might be valid)
        return date_string
        
    except Exception as e:
        logger.warning(f"Failed to parse date '{date_string}': {e}")
        return datetime.now().strftime('%Y-%m-%d')

def create_app():
    app = Flask(__name__)
    # Allow CORS from React frontend
    CORS(app, origins=["http://localhost:3000", "http://127.0.0.1:3000", "http://localhost:3001", "http://127.0.0.1:3001", "http://localhost:3002", "http://127.0.0.1:3002"])

    # Always resolve paths relative to the project root
    PROJECT_ROOT = Path(__file__).parent.parent.parent.resolve()

    # Configure upload folder
    UPLOAD_FOLDER = 'uploads'
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)

    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
    app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

    ALLOWED_EXTENSIONS = {'pdf'}

    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

    @app.route('/api/upload_pdf', methods=['POST'])
    def upload_pdf():
        """Handle PDF upload and extraction using the new extractor2"""
        try:
            # Check if files were uploaded
            if 'files[]' not in request.files:
                # Fallback to single file for backward compatibility
                if 'file' in request.files:
                    files = [request.files['file']]
                else:
                    return jsonify({'success': False, 'error': 'No files provided'}), 400
            else:
                # Multiple files uploaded
                files = request.files.getlist('files[]')
            
            if not files or all(file.filename == '' for file in files):
                return jsonify({'success': False, 'error': 'No files selected'}), 400
            
            # Process each file
            results = []
            total_files = len(files)
            successful_uploads = 0
            
            for file_index, file in enumerate(files):
                if file and allowed_file(file.filename):
                    try:
                        logger.info(f"Processing file {file_index + 1}/{total_files}: {file.filename}")
                        
                        # Secure the filename
                        filename = secure_filename(file.filename)
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        
                        # Save the file
                        file.save(filepath)
                        
                        # Process the PDF
                        file_result = process_single_pdf(filepath, filename, file_index + 1, total_files)
                        
                        if file_result['success']:
                            successful_uploads += 1
                            results.append(file_result)
                        else:
                            results.append(file_result)
                            
                    except Exception as e:
                        logger.error(f"Failed to process file {file.filename}: {e}")
                        results.append({
                            'success': False,
                            'filename': file.filename,
                            'error': f'Processing failed: {str(e)}'
                        })
                else:
                    results.append({
                        'success': False,
                        'filename': file.filename if file else 'unknown',
                        'error': 'Invalid file type. Only PDF files are allowed.'
                    })
            
            # Prepare response data
            response_data = {
                'success': successful_uploads > 0,
                'total_files': total_files,
                'successful_uploads': successful_uploads,
                'failed_uploads': total_files - successful_uploads,
                'results': results,
                'summary': {
                    'total_processed': total_files,
                    'successful': successful_uploads,
                    'failed': total_files - successful_uploads,
                    'success_rate': f"{(successful_uploads/total_files)*100:.1f}%" if total_files > 0 else "0%"
                }
            }
            
            logger.info(f"Batch upload completed: {successful_uploads}/{total_files} successful")
            return jsonify(response_data)
            
        except Exception as e:
            logger.error(f"Server error: {e}")
            return jsonify({'success': False, 'error': f'Server error: {str(e)}'}), 500

    @app.route('/api/upload_multiple_pdfs', methods=['POST'])
    def upload_multiple_pdfs():
        """Handle multiple PDF uploads in batch"""
        return upload_pdf()  # Reuse the same logic
    
    @app.route('/api/batch_status', methods=['GET'])
    def get_batch_status():
        """Get status of batch operations"""
        try:
            csv_path = os.path.join('data', 'extracted_data.csv')
            if not os.path.exists(csv_path):
                return jsonify({
                    'success': True,
                    'batch_status': 'no_data',
                    'message': 'No batch operations performed yet'
                })
            
            # Read CSV data
            df = pd.read_csv(csv_path)
            data = df.to_dict('records')
            
            # Group by date to show batch results
            from collections import defaultdict
            date_groups = defaultdict(list)
            for row in data:
                date_groups[row['DATE']].append(row)
            
            batch_summary = []
            for date, rows in date_groups.items():
                batch_summary.append({
                    'date': date,
                    'files_processed': len(rows),
                    'sample_data': rows[0] if rows else {}
                })
            
            return jsonify({
                'success': True,
                'batch_status': 'completed',
                'total_batches': len(batch_summary),
                'total_files': len(data),
                'batch_summary': batch_summary
            })
            
        except Exception as e:
            logger.error(f"Error getting batch status: {e}")
            return jsonify({
                'success': False,
                'error': f'Failed to get batch status: {str(e)}'
            }), 500

    def process_single_pdf(filepath: str, filename: str, file_index: int, total_files: int) -> dict:
        """Process a single PDF file and return the result"""
        try:
            logger.info(f"Processing PDF: {filepath}")
            
            # Configure Gemini API (you can move this to environment variables)
            gemini_api_key = os.getenv("GEMINI_API_KEY", "AIzaSyCBPEa1qNq_wT-4ehE5Y7KL9kP2-D1N0NM")
            try:
                configure_gemini(gemini_api_key)
                logger.info("✅ Gemini API configured successfully")
            except Exception as e:
                logger.warning(f"⚠️ Gemini API configuration failed: {e}")
            
            # Find headings in PDF
            logger.info("Searching for financial table headings...")
            results = find_headings_in_pdf(
                filepath,
                DEFAULT_TARGET_TITLES,
                exact_relaxed=True,
                include_terms=["main market"],
                exclude_terms=["nomu"]
            )
            
            # Get unique pages found
            unique_pages = sorted({r.page_number for r in results})
            logger.info(f"Found headings on pages: {unique_pages}")
            
            # Export PNGs for evidence
            screenshots_dir = os.path.join('output', 'screenshots')
            os.makedirs(screenshots_dir, exist_ok=True)
            
            screenshot_paths = []
            if unique_pages:
                try:
                    screenshot_paths = export_pages_to_png(filepath, unique_pages, screenshots_dir, scale=2.0)
                    logger.info(f"Exported screenshots: {screenshot_paths}")
                except Exception as e:
                    logger.warning(f"Screenshot export failed: {e}")
            
            # Extract financial metrics using Gemini
            extracted_metrics = None
            ownership_metrics = None
            report_date = None
            
            if unique_pages:
                try:
                    # Extract from Value Traded page (first found page)
                    page_to_extract = unique_pages[0]
                    logger.info(f"🤖 Extracting from page {page_to_extract} (Value Traded)...")
                    extracted_metrics = extract_financial_metrics_with_gemini(filepath, page_to_extract, "value_traded")
                    
                    # Try to extract from Ownership page if it exists
                    ownership_pages = [p for p in unique_pages if p != page_to_extract]
                    if ownership_pages:
                        ownership_page = ownership_pages[0]
                        logger.info(f"🤖 Extracting from page {ownership_page} (Ownership Value)...")
                        ownership_metrics = extract_financial_metrics_with_gemini(filepath, ownership_page, "ownership_value")
                    
                    # Extract report date from either page
                    if extracted_metrics and extracted_metrics.get('REPORT_DATE') and extracted_metrics['REPORT_DATE'] != "NOT_FOUND":
                        report_date = parse_report_date(extracted_metrics['REPORT_DATE'])
                        logger.info(f"📅 Extracted date from Value Traded page: {report_date}")
                    elif ownership_metrics and ownership_metrics.get('REPORT_DATE') and ownership_metrics['REPORT_DATE'] != "NOT_FOUND":
                        report_date = parse_report_date(ownership_metrics['REPORT_DATE'])
                        logger.info(f"📅 Extracted date from Ownership page: {report_date}")
                    
                    # If no date found, try to extract from filename
                    if not report_date or report_date == "NOT_FOUND":
                        # Extract date from filename (e.g., "WeeklyTradingandOwnershipByNationalityReport14-11-2024-2.pdf")
                        import re
                        filename_match = re.search(r'(\d{1,2})-(\d{1,2})-(\d{4})', filename)
                        if filename_match:
                            day, month, year = filename_match.groups()
                            report_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
                            logger.info(f"📅 Extracted date from filename: {report_date}")
                    
                    # Fallback to current date if still no date found
                    if not report_date or report_date == "NOT_FOUND":
                        report_date = datetime.now().strftime('%Y-%m-%d')
                        logger.info(f"📅 Using current date as fallback: {report_date}")
                        
                except Exception as e:
                    logger.error(f"Gemini extraction failed: {e}")
                    report_date = datetime.now().strftime('%Y-%m-%d')
            
            # Prepare extracted data in the format expected by frontend
            extracted_data = {
                'DATE': report_date,
                'Saudi_ValueTraded_Individuals': extracted_metrics.get('INDIVIDUALS') if extracted_metrics else None,
                'Saudi_ValueTraded_Institutions': extracted_metrics.get('INSTITUTIONS') if extracted_metrics else None,
                'GCC_ValueTraded_Total': extracted_metrics.get('GCC') if extracted_metrics else None,
                'Foreign_ValueTraded_Total': extracted_metrics.get('FOREIGN') if extracted_metrics else None,
                'Ownership Value': None,  # Placeholder
                'Saudi_OwnershipValue_Individuals': ownership_metrics.get('INDIVIDUALS') if ownership_metrics else None,
                'Saudi_OwnershipValue_Institutions': ownership_metrics.get('INSTITUTIONS') if ownership_metrics else None,
                'GCC_OwnershipValue_Total': ownership_metrics.get('GCC') if ownership_metrics else None,
                'Foreign_OwnershipValue_Total': ownership_metrics.get('FOREIGN') if ownership_metrics else None,
            }
            
            # Clean up None values
            for key in extracted_data:
                if extracted_data[key] is None:
                    extracted_data[key] = ""
            
            # Debug logging
            logger.info(f"Extracted data: {extracted_data}")
            
            # Save to CSV
            csv_path = os.path.join('data', 'extracted_data.csv')
            append_extraction_to_csv(extracted_data, csv_path)
            
            # Prepare response data for this file
            file_response = {
                'success': True,
                'file_index': file_index,
                'total_files': total_files,
                'data': extracted_data,
                'screenshot_paths': screenshot_paths,
                'filename': filename,
                'filepath': filepath,
                'pages_found': unique_pages,
                'extraction_method': 'extractor2_gemini',
                'value_traded_metrics': extracted_metrics,
                'ownership_metrics': ownership_metrics
            }
            
            logger.info(f"File {file_index}/{total_files} processed successfully: {filename}")
            return file_response
            
        except Exception as e:
            logger.error(f"Failed to process PDF {filename}: {e}")
            return {
                'success': False,
                'file_index': file_index,
                'total_files': total_files,
                'filename': filename,
                'error': f'Processing failed: {str(e)}'
            }

    def append_extraction_to_csv(row, csv_path):
        """
        Append a single extraction result (dict) as a row to a CSV file.
        If the file does not exist, create it with the correct header.
        """
        import csv
        
        CSV_HEADER = [
            'DATE',
            'Saudi_ValueTraded_Individuals',
            'Saudi_ValueTraded_Institutions',
            'GCC_ValueTraded_Total',
            'Foreign_ValueTraded_Total',
            'Ownership Value',
            'Saudi_OwnershipValue_Individuals',
            'Saudi_OwnershipValue_Institutions',
            'GCC_OwnershipValue_Total',
            'Foreign_OwnershipValue_Total',
        ]
        
        file_exists = os.path.isfile(csv_path)
        with open(csv_path, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=CSV_HEADER)
            if not file_exists:
                writer.writeheader()
            # Only write the fields in the header, fill missing with empty string
            row_to_write = {k: row.get(k, '') for k in CSV_HEADER}
            writer.writerow(row_to_write)

    @app.route('/api/get_extracted_data', methods=['GET'])
    def get_extracted_data():
        """Get all extracted data from CSV file"""
        try:
            csv_path = os.path.join('data', 'extracted_data.csv')
            if not os.path.exists(csv_path):
                return jsonify({
                    'success': True,
                    'data': [],
                    'message': 'No data extracted yet'
                })
            
            # Read CSV data
            df = pd.read_csv(csv_path)
            data = df.to_dict('records')
            
            return jsonify({
                'success': True,
                'data': data,
                'count': len(data)
            })
            
        except Exception as e:
            logger.error(f"Error reading extracted data: {e}")
            return jsonify({
                'success': False,
                'error': f'Failed to read data: {str(e)}'
            }), 500

    @app.route('/api/get_screenshots', methods=['GET'])
    def get_screenshots():
        """Get list of available screenshots"""
        try:
            screenshots_dir = os.path.join('output', 'screenshots')
            if not os.path.exists(screenshots_dir):
                return jsonify({
                    'success': True,
                    'screenshots': []
                })
            
            screenshots = []
            for filename in os.listdir(screenshots_dir):
                if filename.endswith('.png'):
                    screenshots.append({
                        'filename': filename,
                        'path': f'/api/screenshot/{filename}',
                        'size': os.path.getsize(os.path.join(screenshots_dir, filename))
                    })
            
            return jsonify({
                'success': True,
                'screenshots': screenshots
            })
            
        except Exception as e:
            logger.error(f"Error getting screenshots: {e}")
            return jsonify({
                'success': False,
                'error': f'Failed to get screenshots: {str(e)}'
            }), 500

    @app.route('/api/screenshot/<filename>', methods=['GET'])
    def get_screenshot(filename):
        """Serve a screenshot file"""
        try:
            screenshots_dir = os.path.join('output', 'screenshots')
            filepath = os.path.join(screenshots_dir, filename)
            
            if not os.path.exists(filepath):
                return jsonify({'success': False, 'error': 'Screenshot not found'}), 404
            
            return send_file(filepath, mimetype='image/png')
            
        except Exception as e:
            logger.error(f"Error serving screenshot: {e}")
            return jsonify({
                'success': False,
                'error': f'Failed to serve screenshot: {str(e)}'
            }), 500

    @app.route('/api/refresh', methods=['POST'])
    def refresh_data():
        """
        Simple refresh endpoint for the new PDF upload system
        """
        try:
            logger.info("Data refresh requested...")
            return jsonify({
                "status": "success", 
                "message": "System refreshed successfully."
            }), 200
            
        except Exception as e:
            logger.error(f"Error during data refresh: {e}")
            return jsonify({
                "status": "error", 
                "message": f"Refresh failed: {str(e)}"
            }), 500

    @app.route('/api/clear_data', methods=['POST'])
    def clear_data():
        """
        Clear all extracted data from CSV file
        """
        try:
            csv_path = os.path.join('data', 'extracted_data.csv')
            if os.path.exists(csv_path):
                # Remove the CSV file
                os.remove(csv_path)
                logger.info("CSV file cleared successfully")
            else:
                logger.info("CSV file does not exist, nothing to clear")
            
            return jsonify({
                "status": "success", 
                "message": "Data cleared successfully."
            }), 200
            
        except Exception as e:
            logger.error(f"Error clearing data: {e}")
            return jsonify({
                "status": "error", 
                "message": f"Clear failed: {str(e)}"
            }), 500

    @app.route('/api/health', methods=['GET'])
    def health_check():
        """Health check endpoint"""
        return jsonify({
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "extractor": "extractor2_gemini"
        })

    @app.route('/api/list_pdfs', methods=['GET'])
    def list_pdfs():
        """
        List all stored PDF files
        """
        try:
            upload_folder = app.config['UPLOAD_FOLDER']
            if not os.path.exists(upload_folder):
                return jsonify({"pdfs": [], "message": "No uploads folder found"})
            
            pdf_files = []
            for filename in os.listdir(upload_folder):
                if filename.lower().endswith('.pdf'):
                    filepath = os.path.join(upload_folder, filename)
                    file_stats = os.stat(filepath)
                    pdf_files.append({
                        'filename': filename,
                        'filepath': filepath,
                        'size_bytes': file_stats.st_size,
                        'upload_date': datetime.fromtimestamp(file_stats.st_mtime).isoformat()
                    })
            
            # Sort by upload date (newest first)
            pdf_files.sort(key=lambda x: x['upload_date'], reverse=True)
            
            return jsonify({
                "status": "success",
                "pdfs": pdf_files,
                "count": len(pdf_files)
            })
            
        except Exception as e:
            logger.error(f"Error listing PDFs: {e}")
            return jsonify({
                "status": "error", 
                "message": f"Failed to list PDFs: {str(e)}"
            }), 500



    @app.route('/api/export_pdf_data', methods=['GET'])
    def export_pdf_data():
        """
        Export PDF extraction data to Excel file with elegant formatting
        """
        try:
            import pandas as pd
            from pathlib import Path
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            from io import BytesIO
            
            # Path to our extracted data CSV
            csv_path = Path("data/extracted_data.csv")
            
            if not csv_path.exists():
                return jsonify({"error": "No PDF extraction data found. Please upload some PDFs first."}), 404
            
            # Read the CSV data and filter out empty rows
            data = pd.read_csv(csv_path)
            # Remove rows where all data columns are empty or null
            data_columns = ['Saudi_ValueTraded_Individuals', 'Saudi_ValueTraded_Institutions', 
                           'GCC_ValueTraded_Total', 'Foreign_ValueTraded_Total',
                           'Saudi_OwnershipValue_Individuals', 'Saudi_OwnershipValue_Institutions',
                           'GCC_OwnershipValue_Total', 'Foreign_OwnershipValue_Total']
            data = data.dropna(subset=data_columns, how='all')
            # Remove rows where all values are empty strings
            data = data[~(data[data_columns] == '').all(axis=1)]
            
            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF_Extraction_Data"
            
            # Define styles
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            subheader_font = Font(name='Calibri', size=11, bold=True, color='495057')
            data_font = Font(name='Calibri', size=10)
            
            green_fill = PatternFill(start_color='1E6641', end_color='1E6641', fill_type='solid')
            light_gray_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            thick_right_border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            column_widths = {
                'A': 20,  # DATE
                'B': 18,  # Saudi_ValueTraded_Individuals
                'C': 18,  # Saudi_ValueTraded_Institutions
                'D': 15,  # GCC_ValueTraded_Total
                'E': 15,  # Foreign_ValueTraded_Total
                'F': 18,  # Saudi_OwnershipValue_Individuals
                'G': 18,  # Saudi_OwnershipValue_Institutions
                'H': 15,  # GCC_OwnershipValue_Total
                'I': 15,  # Foreign_OwnershipValue_Total
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Create hierarchical header structure
            # Row 1: Main headers
            ws.merge_cells('B1:E1')  # Value Traded
            ws.merge_cells('F1:I1')  # Ownership Value
            
            ws['A1'] = 'DATE'
            ws['B1'] = 'Value Traded'
            ws['F1'] = 'Ownership Value'
            
            # Apply styling to main headers
            for cell in [ws['A1'], ws['B1'], ws['F1']]:
                cell.font = header_font
                cell.fill = green_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Row 2: Sub-headers
            ws.merge_cells('B2:C2')  # Saudi
            ws.merge_cells('F2:G2')  # Saudi
            
            ws['A2'] = ''  # Empty for DATE column
            ws['B2'] = 'Saudi'
            ws['D2'] = 'GCC'
            ws['E2'] = 'Foreign'
            ws['F2'] = 'Saudi'
            ws['H2'] = 'GCC'
            ws['I2'] = 'Foreign'
            
            # Apply styling to sub-headers
            for cell in [ws['B2'], ws['D2'], ws['E2'], ws['F2'], ws['H2'], ws['I2']]:
                cell.font = subheader_font
                cell.fill = light_gray_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Row 3: Detail headers
            ws['A3'] = ''  # Empty for DATE column
            ws['B3'] = 'Individuals'
            ws['C3'] = 'Institutions'
            ws['D3'] = 'Total'
            ws['E3'] = 'Total'
            ws['F3'] = 'Individuals'
            ws['G3'] = 'Institutions'
            ws['H3'] = 'Total'
            ws['I3'] = 'Total'
            
            # Apply styling to detail headers
            for cell in [ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'], ws['G3'], ws['H3'], ws['I3']]:
                cell.font = Font(name='Calibri', size=10, bold=True, color='6C757D')
                cell.fill = white_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Add data rows starting from row 4
            for idx, row in data.iterrows():
                row_num = idx + 4
                
                # DATE column
                ws[f'A{row_num}'] = row.get('DATE', '')
                ws[f'A{row_num}'].font = data_font
                ws[f'A{row_num}'].fill = white_fill if row_num % 2 == 0 else light_gray_fill
                ws[f'A{row_num}'].border = thick_right_border
                ws[f'A{row_num}'].alignment = center_alignment
                
                # Data columns
                columns = [
                    ('B', 'Saudi_ValueTraded_Individuals'),
                    ('C', 'Saudi_ValueTraded_Institutions'),
                    ('D', 'GCC_ValueTraded_Total'),
                    ('E', 'Foreign_ValueTraded_Total'),
                    ('F', 'Saudi_OwnershipValue_Individuals'),
                    ('G', 'Saudi_OwnershipValue_Institutions'),
                    ('H', 'GCC_OwnershipValue_Total'),
                    ('I', 'Foreign_OwnershipValue_Total'),
                ]
                
                for col, field in columns:
                    value = row.get(field, '')
                    cell = ws[f'{col}{row_num}']
                    cell.value = value
                    cell.font = data_font
                    cell.fill = white_fill if row_num % 2 == 0 else light_gray_fill
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    
                    # Format numbers with commas and handle negative values
                    if isinstance(value, (int, float)) and value != '':
                        if value < 0:
                            cell.number_format = '#,##0'
                            cell.font = Font(name='Calibri', size=10, color='D32F2F')
                        else:
                            cell.number_format = '#,##0'
                            cell.font = Font(name='Calibri', size=10, color='2E7D32')
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return the file for download
            return send_file(
                BytesIO(output.read()),
                as_attachment=True,
                download_name=f"pdf_extraction_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
                
        except Exception as e:
            logger.error(f"Error exporting PDF data to Excel: {e}")
            return jsonify({"error": f"Export failed: {str(e)}"}), 500

    @app.route('/api/export_current_table', methods=['POST'])
    def export_current_table():
        """
        Export current table data to Excel file with elegant formatting
        """
        try:
            import pandas as pd
            from pathlib import Path
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            from io import BytesIO
            
            # Get the data from the request
            request_data = request.get_json()
            if not request_data or 'data' not in request_data:
                return jsonify({"error": "No data provided"}), 400
            
            data = pd.DataFrame(request_data['data'])
            
            if data.empty:
                return jsonify({"error": "No data to export"}), 400
            
            # Create a new workbook and select the active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "PDF_Extraction_Data"
            
            # Define styles
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            subheader_font = Font(name='Calibri', size=11, bold=True, color='495057')
            data_font = Font(name='Calibri', size=10)
            
            green_fill = PatternFill(start_color='1E6641', end_color='1E6641', fill_type='solid')
            light_gray_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            thick_right_border = Border(
                left=Side(style='thin'),
                right=Side(style='thick'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            # Set column widths
            column_widths = {
                'A': 20,  # DATE
                'B': 18,  # Saudi_ValueTraded_Individuals
                'C': 18,  # Saudi_ValueTraded_Institutions
                'D': 15,  # GCC_ValueTraded_Total
                'E': 15,  # Foreign_ValueTraded_Total
                'F': 18,  # Saudi_OwnershipValue_Individuals
                'G': 18,  # Saudi_OwnershipValue_Institutions
                'H': 15,  # GCC_OwnershipValue_Total
                'I': 15,  # Foreign_OwnershipValue_Total
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Create hierarchical header structure
            # Row 1: Main headers
            ws.merge_cells('B1:E1')  # Value Traded
            ws.merge_cells('F1:I1')  # Ownership Value
            
            ws['A1'] = 'DATE'
            ws['B1'] = 'Value Traded'
            ws['F1'] = 'Ownership Value'
            
            # Apply styling to main headers
            for cell in [ws['A1'], ws['B1'], ws['F1']]:
                cell.font = header_font
                cell.fill = green_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Row 2: Sub-headers
            ws.merge_cells('B2:C2')  # Saudi
            ws.merge_cells('F2:G2')  # Saudi
            
            ws['A2'] = ''  # Empty for DATE column
            ws['B2'] = 'Saudi'
            ws['D2'] = 'GCC'
            ws['E2'] = 'Foreign'
            ws['F2'] = 'Saudi'
            ws['H2'] = 'GCC'
            ws['I2'] = 'Foreign'
            
            # Apply styling to sub-headers
            for cell in [ws['B2'], ws['D2'], ws['E2'], ws['F2'], ws['H2'], ws['I2']]:
                cell.font = subheader_font
                cell.fill = light_gray_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Row 3: Detail headers
            ws['A3'] = ''  # Empty for DATE column
            ws['B3'] = 'Individuals'
            ws['C3'] = 'Institutions'
            ws['D3'] = 'Total'
            ws['E3'] = 'Total'
            ws['F3'] = 'Individuals'
            ws['G3'] = 'Institutions'
            ws['H3'] = 'Total'
            ws['I3'] = 'Total'
            
            # Apply styling to detail headers
            for cell in [ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'], ws['G3'], ws['H3'], ws['I3']]:
                cell.font = Font(name='Calibri', size=10, bold=True, color='6C757D')
                cell.fill = white_fill
                cell.border = thin_border
                cell.alignment = center_alignment
            
            # Add data rows starting from row 4
            for idx, row in data.iterrows():
                row_num = idx + 4
                
                # DATE column
                ws[f'A{row_num}'] = row.get('DATE', '')
                ws[f'A{row_num}'].font = data_font
                ws[f'A{row_num}'].fill = white_fill if row_num % 2 == 0 else light_gray_fill
                ws[f'A{row_num}'].border = thick_right_border
                ws[f'A{row_num}'].alignment = center_alignment
                
                # Data columns
                columns = [
                    ('B', 'Saudi_ValueTraded_Individuals'),
                    ('C', 'Saudi_ValueTraded_Institutions'),
                    ('D', 'GCC_ValueTraded_Total'),
                    ('E', 'Foreign_ValueTraded_Total'),
                    ('F', 'Saudi_OwnershipValue_Individuals'),
                    ('G', 'Saudi_OwnershipValue_Institutions'),
                    ('H', 'GCC_OwnershipValue_Total'),
                    ('I', 'Foreign_OwnershipValue_Total'),
                ]
                
                for col, field in columns:
                    value = row.get(field, '')
                    cell = ws[f'{col}{row_num}']
                    cell.value = value
                    cell.font = data_font
                    cell.fill = white_fill if row_num % 2 == 0 else light_gray_fill
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    
                    # Format numbers with commas and handle negative values
                    if isinstance(value, (int, float)) and value != '':
                        if value < 0:
                            cell.number_format = '#,##0'
                            cell.font = Font(name='Calibri', size=10, color='D32F2F')
                        else:
                            cell.number_format = '#,##0'
                            cell.font = Font(name='Calibri', size=10, color='2E7D32')
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Return the file for download
            return send_file(
                BytesIO(output.read()),
                as_attachment=True,
                download_name=f"current_table_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
                
        except Exception as e:
            logger.error(f"Error exporting current table to Excel: {e}")
            return jsonify({"error": f"Export failed: {str(e)}"}), 500



    return app

if __name__ == '__main__':
    app = create_app()
    
    # Ensure screenshots directory exists (use absolute path)
    PROJECT_ROOT = Path(__file__).parent.parent.parent.resolve()
    (PROJECT_ROOT / "output/screenshots").mkdir(parents=True, exist_ok=True)
    
    print(f"Starting Evidence API server...")
    print(f"Screenshots directory: {PROJECT_ROOT / 'output/screenshots'}")
    print(f"API will be available at: http://localhost:5003")
    
    app.run(debug=True, host='0.0.0.0', port=5003) 